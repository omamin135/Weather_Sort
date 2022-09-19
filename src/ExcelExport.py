from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl.utils.cell

from datetime import datetime as dt

from user_preference.UserPreference import requirements, preference
from Process import indexdt







class ExportWorkbook:
    
    def __init__(self, fileName):
        #set up workbook
        self.wb = Workbook()
        self.fileName = fileName

        #initalizse worksheet   
        self.ws = self.wb.active
        self.ws.title = "Weather"

    def saveWorkbook(self):
        self.wb.save(self.fileName)

    def pullPreferenceData(self):
  
        time = [requirements["time"]["start"]]
        time.append(requirements["time"]["end"])
        time.append(preference["time"]["start"])
        time.append(preference["time"]["end"])

        temp = [requirements["temp"]["min"]]
        temp.append(requirements["temp"]["max"])
        temp.append(preference["temp"]["max"])
        temp.append(preference["temp"]["min"])

        humidity = [requirements["humidity"]["min"]]
        humidity.append(requirements["humidity"]["max"])
        humidity.append(preference["humidity"]["max"])
        humidity.append(preference["humidity"]["min"])

        pressure = [requirements["pressure"]["min"]]
        pressure.append(requirements["pressure"]["max"])
        pressure.append(preference["pressure"]["max"])
        pressure.append(preference["pressure"]["min"])

        speed = [requirements["wind_speed"]["min"]]
        speed.append(requirements["wind_speed"]["max"])
        speed.append(preference["wind_speed"]["max"])
        speed.append(preference["wind_speed"]["min"])

        precp = [requirements["precipitation%"]["min"]]
        precp.append(requirements["precipitation%"]["max"])
        precp.append(preference["precipitation%"]["max"])
        precp.append(preference["precipitation%"]["min"])

        vis = [requirements["visibility"]["min"]]
        vis.append(requirements["visibility"]["max"])
        vis.append(preference["visibility"]["max"])
        vis.append(preference["visibility"]["min"])

        data = []
        data.append(time)
        data.append(temp)
        data.append(humidity)
        data.append(pressure)
        data.append(speed)
        data.append(precp)
        data.append(vis)

        return data

    def exportPreferenceData(self, data):
        '''
        export user preference data to spreadsheet
        '''
        
        # export category row titles
        self.ws.cell(row=4, column=1).value = "Time"
        self.ws.cell(row=5, column=1).value = "Temperature"
        self.ws.cell(row=6, column=1).value = "Humidity"
        self.ws.cell(row=7, column=1).value = "Pressure"
        self.ws.cell(row=8, column=1).value = "Wind Speed"
        self.ws.cell(row=9, column=1).value = "Precipitation %"
        self.ws.cell(row=10, column=1).value = "Visibility"

        # export colum headers
        self.ws.cell(row=2, column=2).value = "User Requirements"
        self.ws.cell(row=2, column=4).value = "User Preferences"
        self.ws.cell(row=3, column=2).value = "min"
        self.ws.cell(row=3, column=3).value = "max"
        self.ws.cell(row=3, column=4).value = "min"
        self.ws.cell(row=3, column=5).value = "max"

      
        # export user preference data
        for r in range(len(data)):
            for c in range(len(data[0])):
                if (data[r][c] == None):
                    data[r][c] = "-"
                self.ws.cell(row=r+4, column=c+2).value = data[r][c]
        
    def exportDtData(self, filteredData, sortedData):
        
        self.ws.cell(row=3, column=6).value = r"% Fit:"

        for i in range(len(sortedData)-1, -1, -1):

            index = indexdt(filteredData, sortedData[i][0])
            
            percent = sortedData[i][1] * 100

            l = len(sortedData) - i

            dateStr = filteredData[index]["dt_txt"]
            date = dt.strptime(dateStr, "%Y-%m-%d %H:%M:%S")

            self.ws.cell(row=2, column=l+6).value = date.date()
            self.ws.cell(row=3, column=l+6).value = percent / 100
            self.ws.cell(row=4, column=l+6).value = date.time()

            self.ws.cell(row=5, column=l+6).value = filteredData[index]["main"]["temp"]
            self.ws.cell(row=6, column=l+6).value = filteredData[index]["main"]["humidity"]
            self.ws.cell(row=7, column=l+6).value = filteredData[index]["main"]["pressure"]
            self.ws.cell(row=8, column=l+6).value = filteredData[index]["wind"]["speed"]
            self.ws.cell(row=9, column=l+6).value = filteredData[index]["pop"]
            self.ws.cell(row=10, column=l+6).value = filteredData[index]["visibility"]
            
            
    
    def formatSpreadSheet(self, sortedDataLength):
        
        #merge User Requirements and User Preference cells
        self.ws.merge_cells("B2:C2")
        self.ws.merge_cells("D2:E2")

        #adjust column A width
        self.ws.column_dimensions['A'].width = 15

        #set font and color styles
        fontStyle = Font(bold=True)
        darkGrayStyle = PatternFill(start_color='BFBFBF', 
                                end_color='BFBFBF', 
                                fill_type='solid')
        lightGrayStyle = PatternFill(start_color='D9D9D9', 
                                    end_color='D9D9D9', 
                                    fill_type='solid')
        
        centerAlign = Alignment(horizontal='center')
        rightAlign = Alignment(horizontal="right")
        
        #center align user preference data
        for r in range(4,11):
            for c in range(2,6):
                self.ws.cell(row=r, column=c).alignment = centerAlign

        for i in range(5, 10, 2):
            for r in self.ws[i:i]:
                r.fill = lightGrayStyle

        #bold and color title
        for r in range(4,11):
            #column A
            self.ws.cell(row=r, column=1).font = fontStyle
            self.ws.cell(row=r, column=1).fill = darkGrayStyle
            
        
        for c in range(2,6):
            #User requriement and preference header + max and min
            self.ws.cell(row=3, column=c).font = fontStyle
            self.ws.cell(row=3, column=c).fill = darkGrayStyle
        
            self.ws.cell(row=2, column=c).font = Font(bold=True, size=13)
            self.ws.cell(row=2, column=c).font = Font(bold=True, size=13)
            self.ws.column_dimensions[openpyxl.utils.cell.get_column_letter(c)].width = 10
            self.ws.cell(row=2, column=c).alignment = centerAlign
            self.ws.cell(row=3, column=c).alignment = centerAlign
        

        for i in range(sortedDataLength+1):
            
            self.ws.column_dimensions[openpyxl.utils.cell.get_column_letter(i+6)].width = 12
            self.ws.cell(row=2, column=i+6).font = Font(bold=True, size = 13)
            self.ws.cell(row=2, column=i+6).number_format = "mm-dd-yy"
            self.ws.cell(row=3, column=i+6).number_format = "0.00%"
            self.ws.cell(row=4, column=i+6).number_format = "h:mm AM/PM"

        self.ws.cell(row=3, column=6).font = fontStyle
        self.ws.cell(row=3, column=6).fill = darkGrayStyle
        self.ws.cell(row=3, column=6).alignment = rightAlign

       
        for r in range(2, 11):
            for c in range(1, sortedDataLength+7):
                self.ws.cell(row=r, column=c).border = Border(left=Side(style='thin'), 
                                                            right=Side(style='thin'), 
                                                            top=Side(style='thin'), 
                                                            bottom=Side(style='thin'))

     

        

      

     
                

        